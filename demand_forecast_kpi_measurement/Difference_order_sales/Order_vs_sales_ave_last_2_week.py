import calendar
import datetime
import os
from datetime import timedelta
from pyspark.sql import SQLContext
from pyspark.sql import SparkSession
from openpyxl import Workbook
from os.path import abspath
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt

warehouse_location = abspath('spark-warehouse')
os.environ["PYSPARK_SUBMIT_ARGS"] = '--jars /data/jupyter/kudu-spark2_2.11-1.8.0.jar pyspark-shell'

spark = SparkSession.builder \
    .appName("Generate order forecast file") \
    .config("spark.sql.warehouse.dir", warehouse_location) \
    .config("spark.driver.memory", '8g') \
    .config("spark.executor.memory", '6g') \
    .config("spark.num.executors", '14') \
    .config("hive.exec.compress.output", 'false') \
    .config("spark.sql.crossJoin.enabled", 'true') \
    .config("spark.sql.autoBroadcastJoinThreshold", '-1') \
    .enableHiveSupport() \
    .getOrCreate()

sc = spark.sparkContext

sqlc = SQLContext(sc)

print ("start:{}".format(datetime.datetime.now()))

### 1. Parameter Setting(output path/rundate)

output_path = "Carrefour-China-Supply-Chain-Forecast/demand_forecast_kpi_measurement/Order_vs_actual_sales_last_two_week"
run_date = datetime.datetime.strptime("20190826",'%Y%m%d').date()
last2week_start_date = (run_date - timedelta(weeks=2)).strftime("%Y%m%d")
w1_start_date = run_date
w2_start_date = run_date + timedelta(weeks=1)
w3_start_date = run_date + timedelta(weeks=2)
w4_start_date = run_date + timedelta(weeks=3)
w5_start_date = run_date + timedelta(weeks=4)
w6_start_date = run_date + timedelta(weeks=5)
w7_start_date = run_date + timedelta(weeks=6)
w8_start_date = run_date + timedelta(weeks=7)
w9_start_date = run_date + timedelta(weeks=8)

print ("Step 1 Ends:{}".format(datetime.datetime.now()))

### 2. Read item-order by supplier by week

#### 2.1 item-order table

xdock_orders_sql = """
    select
        w1.holding_code,
        w1.primary_barcode,
        w1.dept_code,
        w1.item_code,
        w1.sub_code,
        w1.item_name_local,
        w1.item_name_english,
        w1.order_qty as w1_order_qty,
        w1.dm_qty as w1_dm_qty,
        w2.order_qty as w2_order_qty,
        w2.dm_qty as w2_dm_qty,
        w3.order_qty as w3_order_qty,
        w3.dm_qty as w3_dm_qty,
        w4.order_qty as w4_order_qty,
        w4.dm_qty as w4_dm_qty,
        w5.order_qty as w5_order_qty,
        w5.dm_qty as w5_dm_qty,
        w6.order_qty as w6_order_qty,
        w6.dm_qty as w6_dm_qty,
        w7.order_qty as w7_order_qty,
        w7.dm_qty as w7_dm_qty,
        w8.order_qty as w8_order_qty,
        w8.dm_qty as w8_dm_qty,
        w9.order_qty as w9_order_qty,
        w9.dm_qty as w9_dm_qty
    from
        vartefact.v_forecast_weekly_xdock_order_forecast w1
        join vartefact.v_forecast_weekly_xdock_order_forecast w2 on w1.dept_code = w2.dept_code
        and w1.item_code = w2.item_code
        and w1.sub_code = w2.sub_code
        and w1.week_start_day = '{0}'
        and w2.week_start_day = '{1}'
        join vartefact.v_forecast_weekly_xdock_order_forecast w3 on w1.dept_code = w3.dept_code
        and w1.item_code = w3.item_code
        and w1.sub_code = w3.sub_code
        and w3.week_start_day = '{2}'
        join vartefact.v_forecast_weekly_xdock_order_forecast w4 on w1.dept_code = w4.dept_code
        and w1.item_code = w4.item_code
        and w1.sub_code = w4.sub_code
        and w4.week_start_day = '{3}'
        join vartefact.v_forecast_weekly_xdock_order_forecast w5 on w1.dept_code = w5.dept_code
        and w1.item_code = w5.item_code
        and w1.sub_code = w5.sub_code
        and w5.week_start_day = '{4}'
        join vartefact.v_forecast_weekly_xdock_order_forecast w6 on w1.dept_code = w6.dept_code
        and w1.item_code = w6.item_code
        and w1.sub_code = w6.sub_code
        and w6.week_start_day = '{5}'
        join vartefact.v_forecast_weekly_xdock_order_forecast w7 on w1.dept_code = w7.dept_code
        and w1.item_code = w7.item_code
        and w1.sub_code = w7.sub_code
        and w7.week_start_day = '{6}'
        join vartefact.v_forecast_weekly_xdock_order_forecast w8 on w1.dept_code = w8.dept_code
        and w1.item_code = w8.item_code
        and w1.sub_code = w8.sub_code
        and w8.week_start_day = '{7}'
        join vartefact.v_forecast_weekly_xdock_order_forecast w9 on w1.dept_code = w9.dept_code
        and w1.item_code = w9.item_code
        and w1.sub_code = w9.sub_code
        and w9.week_start_day = '{8}'
        """.replace("\n", " ") \
        .format(
                w1_start_date.strftime("%Y%m%d"),
                w2_start_date.strftime("%Y%m%d"),
                w3_start_date.strftime("%Y%m%d"),
                w4_start_date.strftime("%Y%m%d"),
                w5_start_date.strftime("%Y%m%d"),
                w6_start_date.strftime("%Y%m%d"),
                w7_start_date.strftime("%Y%m%d"),
                w8_start_date.strftime("%Y%m%d"),
                w9_start_date.strftime("%Y%m%d"))

xdock_orders = sqlc.sql(xdock_orders_sql)

xdock_orders_df = xdock_orders.toPandas()

dc_orders_sql = """
    select
        w1.holding_code,
        w1.primary_barcode,
        w1.dept_code,
        w1.item_code,
        w1.sub_code,
        w1.item_name_local,
        w1.item_name_english,
        w1.order_qty as w1_order_qty,
        w1.dm_qty as w1_dm_qty,
        w2.order_qty as w2_order_qty,
        w2.dm_qty as w2_dm_qty,
        w3.order_qty as w3_order_qty,
        w3.dm_qty as w3_dm_qty,
        w4.order_qty as w4_order_qty,
        w4.dm_qty as w4_dm_qty,
        w5.order_qty as w5_order_qty,
        w5.dm_qty as w5_dm_qty,
        w6.order_qty as w6_order_qty,
        w6.dm_qty as w6_dm_qty,
        w7.order_qty as w7_order_qty,
        w7.dm_qty as w7_dm_qty,
        w8.order_qty as w8_order_qty,
        w8.dm_qty as w8_dm_qty,
        w9.order_qty as w9_order_qty,
        w9.dm_qty as w9_dm_qty
    from
        vartefact.v_forecast_weekly_dc_order_forecast w1
        join vartefact.v_forecast_weekly_dc_order_forecast w2 on w1.dept_code = w2.dept_code
        and w1.item_code = w2.item_code
        and w1.sub_code = w2.sub_code
        and w1.week_start_day = '{0}'
        and w2.week_start_day = '{1}'
        join vartefact.v_forecast_weekly_dc_order_forecast w3 on w1.dept_code = w3.dept_code
        and w1.item_code = w3.item_code
        and w1.sub_code = w3.sub_code
        and w3.week_start_day = '{2}'
        join vartefact.v_forecast_weekly_dc_order_forecast w4 on w1.dept_code = w4.dept_code
        and w1.item_code = w4.item_code
        and w1.sub_code = w4.sub_code
        and w4.week_start_day = '{3}'
        join vartefact.v_forecast_weekly_dc_order_forecast w5 on w1.dept_code = w5.dept_code
        and w1.item_code = w5.item_code
        and w1.sub_code = w5.sub_code
        and w5.week_start_day = '{4}'
        join vartefact.v_forecast_weekly_dc_order_forecast w6 on w1.dept_code = w6.dept_code
        and w1.item_code = w6.item_code
        and w1.sub_code = w6.sub_code
        and w6.week_start_day = '{5}'
        join vartefact.v_forecast_weekly_dc_order_forecast w7 on w1.dept_code = w7.dept_code
        and w1.item_code = w7.item_code
        and w1.sub_code = w7.sub_code
        and w7.week_start_day = '{6}'
        join vartefact.v_forecast_weekly_dc_order_forecast w8 on w1.dept_code = w8.dept_code
        and w1.item_code = w8.item_code
        and w1.sub_code = w8.sub_code
        and w8.week_start_day = '{7}'
        join vartefact.v_forecast_weekly_dc_order_forecast w9 on w1.dept_code = w9.dept_code
        and w1.item_code = w9.item_code
        and w1.sub_code = w9.sub_code
        and w9.week_start_day = '{8}'
        """.replace("\n", " ") \
        .format(
                w1_start_date.strftime("%Y%m%d"),
                w2_start_date.strftime("%Y%m%d"),
                w3_start_date.strftime("%Y%m%d"),
                w4_start_date.strftime("%Y%m%d"),
                w5_start_date.strftime("%Y%m%d"),
                w6_start_date.strftime("%Y%m%d"),
                w7_start_date.strftime("%Y%m%d"),
                w8_start_date.strftime("%Y%m%d"),
                w9_start_date.strftime("%Y%m%d"))

dc_orders = sqlc.sql(dc_orders_sql)

dc_orders_df = dc_orders.toPandas()

run_date_str = run_date.strftime("%Y%m%d")

w1_date_str = w1_start_date.strftime("%Y%m%d")
w2_date_str = w2_start_date.strftime("%Y%m%d")
w3_date_str = w3_start_date.strftime("%Y%m%d")
w4_date_str = w4_start_date.strftime("%Y%m%d")
w5_date_str = w5_start_date.strftime("%Y%m%d")
w6_date_str = w6_start_date.strftime("%Y%m%d")
w7_date_str = w7_start_date.strftime("%Y%m%d")
w8_date_str = w8_start_date.strftime("%Y%m%d")
w9_date_str = w9_start_date.strftime("%Y%m%d")

con_holding = "700"
forecast_file = f"Carrefour_Order_Forecast_DC_level_{con_holding}_{run_date_str}.xlsx"

wb = Workbook()
ws = wb.active
ws.append(
    ['Supplier_name','Barcode','Department_code','Item_code',
     'Sub_code','Item_desc_chn','Item_desc_eng',
     f'Week1_{w1_date_str}_Permanent_Box', f'Week1_{w1_date_str}_DM_Box',
     f'Week2_{w2_date_str}_Permanent_Box', f'Week2_{w2_date_str}_DM_Box',
     f'Week3_{w3_date_str}_Permanent_Box', f'Week3_{w3_date_str}_DM_Box',
     f'Week4_{w4_date_str}_Permanent_Box', f'Week4_{w4_date_str}_DM_Box',
     f'Week5_{w5_date_str}_Permanent_Box', f'Week5_{w5_date_str}_DM_Box',
     f'Week6_{w6_date_str}_Permanent_Box', f'Week6_{w6_date_str}_DM_Box',
     f'Week7_{w7_date_str}_Permanent_Box', f'Week7_{w7_date_str}_DM_Box',
     f'Week8_{w8_date_str}_Permanent_Box', f'Week8_{w8_date_str}_DM_Box',
     f'Week9_{w9_date_str}_Permanent_Box', f'Week9_{w9_date_str}_DM_Box '])

for index, row in xdock_orders_df[xdock_orders_df["holding_code"] == "700"].iterrows():
    ws.append(["Unilever Services (Hefei) Co. Ltd.", row.primary_barcode,
             row.dept_code, row.item_code, row.sub_code, row.item_name_local, row.item_name_english,
             row.w1_order_qty, row.w1_dm_qty,row.w2_order_qty, row.w2_dm_qty,
              row.w3_order_qty, row.w3_dm_qty,row.w4_order_qty, row.w4_dm_qty,
              row.w5_order_qty, row.w5_dm_qty,row.w6_order_qty, row.w6_dm_qty,
              row.w7_order_qty, row.w7_dm_qty,row.w8_order_qty, row.w8_dm_qty,
              row.w9_order_qty, row.w9_dm_qty])

for index, row in dc_orders_df[dc_orders_df["holding_code"] == "700"].iterrows():
    ws.append(["Unilever Services (Hefei) Co. Ltd.", row.primary_barcode,
             row.dept_code, row.item_code, row.sub_code, row.item_name_local, row.item_name_english,
             row.w1_order_qty, row.w1_dm_qty,row.w2_order_qty, row.w2_dm_qty,
              row.w3_order_qty, row.w3_dm_qty,row.w4_order_qty, row.w4_dm_qty,
              row.w5_order_qty, row.w5_dm_qty,row.w6_order_qty, row.w6_dm_qty,
              row.w7_order_qty, row.w7_dm_qty,row.w8_order_qty, row.w8_dm_qty,
              row.w9_order_qty, row.w9_dm_qty])

wb.save(output_path + '/' + forecast_file)

con_holding = "693"
forecast_file = f"Carrefour_Order_Forecast_DC_level_{con_holding}_{run_date_str}.xlsx"

wb = Workbook()
ws = wb.active
ws.append(
    ['Supplier_name','Barcode','Department_code','Item_code',
     'Sub_code','Item_desc_chn','Item_desc_eng',
     f'Week1_{w1_date_str}_Permanent_Box', f'Week1_{w1_date_str}_DM_Box',
     f'Week2_{w2_date_str}_Permanent_Box', f'Week2_{w2_date_str}_DM_Box',
     f'Week3_{w3_date_str}_Permanent_Box', f'Week3_{w3_date_str}_DM_Box',
     f'Week4_{w4_date_str}_Permanent_Box', f'Week4_{w4_date_str}_DM_Box',
     f'Week5_{w5_date_str}_Permanent_Box', f'Week5_{w5_date_str}_DM_Box',
     f'Week6_{w6_date_str}_Permanent_Box', f'Week6_{w6_date_str}_DM_Box',
     f'Week7_{w7_date_str}_Permanent_Box', f'Week7_{w7_date_str}_DM_Box',
     f'Week8_{w8_date_str}_Permanent_Box', f'Week8_{w8_date_str}_DM_Box',
     f'Week9_{w9_date_str}_Permanent_Box', f'Week9_{w9_date_str}_DM_Box '])

for index, row in xdock_orders_df[xdock_orders_df["holding_code"] == "693"].iterrows():
    ws.append(["Procter&Gamble (China) Sales Co.,Ltd.", row.primary_barcode,
             row.dept_code, row.item_code, row.sub_code, row.item_name_local, row.item_name_english,
             row.w1_order_qty, row.w1_dm_qty,row.w2_order_qty, row.w2_dm_qty,
              row.w3_order_qty, row.w3_dm_qty,row.w4_order_qty, row.w4_dm_qty,
              row.w5_order_qty, row.w5_dm_qty,row.w6_order_qty, row.w6_dm_qty,
              row.w7_order_qty, row.w7_dm_qty,row.w8_order_qty, row.w8_dm_qty,
              row.w9_order_qty, row.w9_dm_qty])

for index, row in dc_orders_df[dc_orders_df["holding_code"] == "693"].iterrows():
    ws.append(["Procter&Gamble (China) Sales Co.,Ltd.", row.primary_barcode,
             row.dept_code, row.item_code, row.sub_code, row.item_name_local, row.item_name_english,
             row.w1_order_qty, row.w1_dm_qty,row.w2_order_qty, row.w2_dm_qty,
              row.w3_order_qty, row.w3_dm_qty,row.w4_order_qty, row.w4_dm_qty,
              row.w5_order_qty, row.w5_dm_qty,row.w6_order_qty, row.w6_dm_qty,
              row.w7_order_qty, row.w7_dm_qty,row.w8_order_qty, row.w8_dm_qty,
              row.w9_order_qty, row.w9_dm_qty])

wb.save(output_path + '/' + forecast_file)

con_holding = "002"
forecast_file = f"Carrefour_Order_Forecast_DC_level_{con_holding}_{run_date_str}.xlsx"

wb = Workbook()
ws = wb.active
ws.append(
    ['Supplier_name','Barcode','Department_code','Item_code',
     'Sub_code','Item_desc_chn','Item_desc_eng',
     f'Week1_{w1_date_str}_Permanent_Box', f'Week1_{w1_date_str}_DM_Box',
     f'Week2_{w2_date_str}_Permanent_Box', f'Week2_{w2_date_str}_DM_Box',
     f'Week3_{w3_date_str}_Permanent_Box', f'Week3_{w3_date_str}_DM_Box',
     f'Week4_{w4_date_str}_Permanent_Box', f'Week4_{w4_date_str}_DM_Box',
     f'Week5_{w5_date_str}_Permanent_Box', f'Week5_{w5_date_str}_DM_Box',
     f'Week6_{w6_date_str}_Permanent_Box', f'Week6_{w6_date_str}_DM_Box',
     f'Week7_{w7_date_str}_Permanent_Box', f'Week7_{w7_date_str}_DM_Box',
     f'Week8_{w8_date_str}_Permanent_Box', f'Week8_{w8_date_str}_DM_Box',
     f'Week9_{w9_date_str}_Permanent_Box', f'Week9_{w9_date_str}_DM_Box '])

for index, row in xdock_orders_df[xdock_orders_df["holding_code"] == "002"].iterrows():
    ws.append(["Shanghai Nestle products Service Co.,Ltd", row.primary_barcode,
             row.dept_code, row.item_code, row.sub_code, row.item_name_local, row.item_name_english,
             row.w1_order_qty, row.w1_dm_qty,row.w2_order_qty, row.w2_dm_qty,
              row.w3_order_qty, row.w3_dm_qty,row.w4_order_qty, row.w4_dm_qty,
              row.w5_order_qty, row.w5_dm_qty,row.w6_order_qty, row.w6_dm_qty,
              row.w7_order_qty, row.w7_dm_qty,row.w8_order_qty, row.w8_dm_qty,
              row.w9_order_qty, row.w9_dm_qty])

for index, row in dc_orders_df[dc_orders_df["holding_code"] == "002"].iterrows():
    ws.append(["Shanghai Nestle products Service Co.,Ltd", row.primary_barcode,
             row.dept_code, row.item_code, row.sub_code, row.item_name_local, row.item_name_english,
             row.w1_order_qty, row.w1_dm_qty,row.w2_order_qty, row.w2_dm_qty,
              row.w3_order_qty, row.w3_dm_qty,row.w4_order_qty, row.w4_dm_qty,
              row.w5_order_qty, row.w5_dm_qty,row.w6_order_qty, row.w6_dm_qty,
              row.w7_order_qty, row.w7_dm_qty,row.w8_order_qty, row.w8_dm_qty,
              row.w9_order_qty, row.w9_dm_qty])

wb.save(output_path + '/' + forecast_file)

df_002 = pd.read_excel(f"item_order/Carrefour_Order_Forecast_DC_level_002_{run_date_str}.xlsx")
df_693 = pd.read_excel(f"item_order/Carrefour_Order_Forecast_DC_level_693_{run_date_str}.xlsx")
df_700 = pd.read_excel(f"item_order/Carrefour_Order_Forecast_DC_level_700_{run_date_str}.xlsx")
df_order = df_002.append(df_693)
df_order = df_order.append(df_700)
df_order['week_1_order'] = df_order[f'Week1_{run_date_str}_Permanent_Box'] + df_order[f'Week1_{run_date_str}_DM_Box']
df_order['Item_code'] = df_order['Item_code'].apply(lambda x:str(x).rjust(6,'0'))
df_order['Sub_code'] = df_order['Sub_code'].apply(lambda x:str(x).rjust(3,'0'))
df_order['Department_code'] = df_order['Department_code'].apply(str)

print ("Step 2.1 Ends:{}".format(datetime.datetime.now()))

#### 2.2 item-sales tale

df_avg_sales_sql = f"""
    with item_list as (
        select
            dept_code,
            item_code,
            sub_code,
            store_code,
            date_key,
            daily_sales_sum
        from vartefact.forecast_sprint4_add_dm_to_daily
        where date_key >= {last2week_start_date}
        and date_key <= {run_date_str}
    ),
    item_list_adding_week as
    (
        select
            a.*,
            b.week_key
        from item_list a
        left join vartefact.lddb_calendar b
        on a.date_key = b.date_key
    ),
    item_week_avg as
    (
        select
            dept_code,
            item_code,
            sub_code,
            week_key,
            sum(daily_sales_sum) as sales_avg_last_two_week
        from item_list_adding_week
        group by dept_code,item_code,sub_code,week_key
    )
    select
        dept_code,
        item_code,
        sub_code,
        avg(sales_avg_last_two_week) as sales_avg_last_two_week
    from item_week_avg
    group by dept_code,item_code,sub_code
"""

df_avg_sales = sqlc.sql(df_avg_sales_sql).toPandas()

print ("Step 2.2 Ends:{}".format(datetime.datetime.now()))

#### 2.3 item pack-box tale

df_pack_sql = """
    with item_list as (
    select
        distinct
        dept_code,
        item_code,
        sub_code,
        qty_per_pack,
        pack_per_box,
        rotation
    from vartefact.v_forecast_inscope_store_item_details
    )
    select *
    from item_list
"""

df_pack_sql = sqlc.sql(df_pack_sql).toPandas()

print ("Step 2.3 Ends:{}".format(datetime.datetime.now()))

### 3. Merge table

df_week_1_order = df_order[['Supplier_name', 'Barcode', 'Department_code', 'Item_code', 'Sub_code',
       'Item_desc_chn', 'Item_desc_eng','week_1_order']]
df_order = pd.merge(df_week_1_order,df_pack_sql,left_on = ['Department_code','Item_code','Sub_code'],
                    right_on= ['dept_code','item_code','sub_code'],how = 'left')
df_order['order_qty'] = df_order['week_1_order'] * df_order['qty_per_pack'] * df_order['pack_per_box']
df_all = pd.merge(df_order,df_avg_sales,left_on = ['Department_code','Item_code','Sub_code'],
         right_on=['dept_code','item_code','sub_code'],how = 'left')
df_all['sales_avg_last_two_week'] = df_all['sales_avg_last_two_week'].apply(float)
df_all['diff'] = np.abs(df_all['order_qty'] - df_all['sales_avg_last_two_week'])
df_plot = df_all[['Supplier_name', 'Barcode', 'Department_code', 'Item_code', 'Sub_code','Item_desc_chn', 'Item_desc_eng','order_qty','sales_avg_last_two_week','diff']]

print ("Step 3 Ends:{}".format(datetime.datetime.now()))

#### 4.1 Output 1:table

df_plot.sort_values(by = 'diff',ascending = False).reset_index().\
to_csv(f"{output_path}/item_list_difference_order_vs_actual_sales_last_two_week.csv",index = False)

#### 4.2 Output 2:Difference distribution Graph

plt.figure(figsize=(16,9))
plt.hist(df_plot[~df_plot['diff'].isnull()]['diff'],bins = 100,color = '#2AB6BF')
plt.title('Distribution for piece difference between order and average sales for last two weeks',fontsize = 15)
plt.xlabel('Count',fontsize = 15)
plt.ylabel('Difference',fontsize = 15)
plt.show()
plt.savefig(f"{output_path}/Difference Distribution.png")

print ("Finally Ends:{}".format(datetime.datetime.now()))
