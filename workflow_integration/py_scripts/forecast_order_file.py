import calendar
import datetime
import os
from datetime import timedelta
from os.path import abspath

import pandas as pd
import numpy as np
from openpyxl import Workbook
from pyspark.sql import SQLContext
from pyspark.sql import SparkSession


def store_order_file_process(date_str, record_folder, output_path,
                             datachecking_output_folder,
                             store_order_filename,
                             store_highvalue_order_filename,
                             xdock_high_volume_order_filename):
    warehouse_location = abspath('spark-warehouse')
    os.environ["PYSPARK_SUBMIT_ARGS"] = '--jars /data/jupyter/kudu-spark2_2.11-1.8.0.jar pyspark-shell'

    spark = SparkSession.builder \
        .appName("Generate store order file") \
        .config("spark.sql.warehouse.dir", warehouse_location) \
        .config("spark.driver.memory", '8g') \
        .config("spark.executor.memory", '8g') \
        .config("spark.num.executors", '8') \
        .config("hive.exec.compress.output", 'false') \
        .enableHiveSupport() \
        .getOrCreate()

    sc = spark.sparkContext

    sqlc = SQLContext(sc)
    # -

    # # On Stock Store

    # +

    run_date = datetime.datetime.strptime(date_str, '%Y%m%d').date()

    stock_date = run_date + timedelta(days=-1)
    # -

    onstock_store_items_sql = """
    SELECT DISTINCT
        store_code,
        dept_code,
        item_code,
        sub_code,
        supplier_code
    FROM
    (
        SELECT
        ord.store_code,
        ord.dept_code,
        ord.supplier_code,
        ord.item_code,
        ord.sub_code
        FROM vartefact.forecast_onstock_orders ord
        WHERE ord.order_day = '{0}'

        UNION

        SELECT
        dm.store_code,
        dm.dept_code,
        dm.dc_supplier_code as supplier_code,
        dm.item_code,
        dm.sub_code
        FROM vartefact.forecast_dm_orders dm
        WHERE dm.first_order_date = '{0}'
        and dm.rotation !='X'
    ) ords
        """.replace("\n", " ").format(run_date.strftime("%Y%m%d"))

    onstock_store_items = sqlc.sql(onstock_store_items_sql)
    onstock_store_items.createOrReplaceTempView('onstock_store_items')

    onstock_store_sql = """
    SELECT osi.store_code,
        osi.dept_code,
        osi.supplier_code,
        osi.item_code,
        osi.sub_code,
        ord.order_qty,
        ord.order_without_pcb,
        ord.delivery_day,
        dm.first_dm_order_qty,
        dm.order_without_pcb as dm_order_qty_without_pcb,
        dm.pcb,
        dm.ppp,
        dm.npp,
        dm.four_weeks_after_dm,
        cast(sl.service_level as DOUBLE) service_level,
        id.qty_per_unit,
        id.order_by,
        id.qty_per_pack,
        id.pack_per_box,
        cast(fpsi.npp as DOUBLE) itm_npp,
        ord.item_id,
        ord.sub_id,
        id.cn_name
    FROM onstock_store_items osi
    LEFT JOIN vartefact.forecast_onstock_orders ord
        ON osi.store_code = ord.store_code
        AND osi.dept_code = ord.dept_code
        AND osi.item_code = ord.item_code
        AND osi.sub_code =  ord.sub_code
        AND ord.order_day = '{0}'
    LEFT JOIN vartefact.forecast_dm_orders dm 
        ON osi.store_code = dm.store_code
        AND osi.dept_code = dm.dept_code
        AND osi.item_code = dm.item_code
        AND osi.sub_code =  dm.sub_code
        AND dm.first_order_date = '{0}'
    LEFT JOIN vartefact.service_level_safety2_vinc sl
        on ord.item_code = sl.item_code
        and ord.sub_code = sl.sub_code
        and ord.dept_code = sl.dept_code
    JOIN vartefact.forecast_store_item_details id 
        ON ord.item_code = id.item_code
        AND ord.sub_code = id.sub_code
        AND ord.dept_code = id.dept_code
        AND ord.store_code = id.store_code
    LEFT JOIN vartefact.forecast_p4cm_store_item fpsi
        on ord.item_code = fpsi.item_code
        and ord.sub_code = fpsi.sub_code
        and ord.dept_code = fpsi.dept_code 
        and ord.store_code = fpsi.store_code 
        and fpsi.date_key = '{1}'
        """.replace("\n", " ").format(run_date.strftime("%Y%m%d"), stock_date.strftime("%Y%m%d"))

    onstock_store_df = sqlc.sql(onstock_store_sql)

    onstock_store = onstock_store_df.toPandas()

    # +
    onstock_store['order_qty'] = onstock_store['order_qty'].fillna(0)

    onstock_store['order_qty_without_pcb'] = onstock_store['order_without_pcb'].fillna(0)

    onstock_store['dm_order_qty'] = onstock_store['first_dm_order_qty'].fillna(0)

    onstock_store['pcb'] = onstock_store['pcb'].fillna(1)

    onstock_store['dm_order_qty_without_pcb'] = onstock_store['dm_order_qty_without_pcb'].fillna(0)

    onstock_store['four_weeks_after_dm'] = onstock_store['four_weeks_after_dm'].fillna(0)

    onstock_store['service_level'] = onstock_store['service_level'].fillna(1)

    onstock_store['total_order'] = np.ceil((onstock_store['order_qty_without_pcb'] + onstock_store['dm_order_qty_without_pcb'])
                                       / onstock_store['qty_per_unit']) * onstock_store['qty_per_unit']

    onstock_store['total_order_in_unit'] = onstock_store['total_order'] / onstock_store['qty_per_unit']

    onstock_store['itm_npp'] = onstock_store['itm_npp'].fillna(1)

    onstock_store['order_value'] = onstock_store['itm_npp'] * onstock_store['total_order']

    onstock_store['single_unit_value'] = onstock_store['itm_npp'] * onstock_store['qty_per_unit']
    # -

    xdock_items_sql = """
    SELECT DISTINCT
        store_code,
        dept_code,
        item_code,
        sub_code,
        supplier_code
    FROM
    (
        SELECT
        ord.store_code,
        ord.dept_code,
        ord.supplier_code,
        ord.item_code,
        ord.sub_code
        FROM vartefact.forecast_xdock_orders ord
        WHERE ord.order_day = '{0}'

        UNION

        SELECT
        dm.store_code,
        dm.dept_code,
        dm.dc_supplier_code as supplier_code,
        dm.item_code,
        dm.sub_code
        FROM vartefact.forecast_dm_orders dm
        WHERE dm.first_order_date > '20191007'
        and dm.first_order_date = '{0}'
        and dm.rotation ='X'
    ) ords
        """.replace("\n", " ").format(run_date.strftime("%Y%m%d"))

    xdock_items = sqlc.sql(xdock_items_sql)
    xdock_items.createOrReplaceTempView('xdock_items')

    xdock_sql = """
    SELECT osi.store_code,
        osi.dept_code,
        osi.supplier_code,
        osi.item_code,
        osi.sub_code,
        ord.order_without_pcb,
        ord.delivery_day,
        dm.order_qty as dm_order_qty,
        dm.order_without_pcb as dm_order_qty_without_pcb,
        dm.ppp,
        dm.npp,
        dm.four_weeks_after_dm,
        cast(sl.service_level as DOUBLE) service_level,
        id.qty_per_unit,
        id.order_by,
        id.qty_per_pack,
        id.pack_per_box,
        cast(fpsi.npp as DOUBLE) itm_npp,
        ord.item_id,
        ord.sub_id,
        id.cn_name
    FROM xdock_items osi
    LEFT JOIN vartefact.forecast_xdock_orders ord
        ON osi.store_code = ord.store_code
        AND osi.dept_code = ord.dept_code
        AND osi.item_code = ord.item_code
        AND osi.sub_code =  ord.sub_code
        AND ord.order_day = '{0}'
    LEFT JOIN vartefact.forecast_dm_orders dm 
        ON osi.store_code = dm.store_code
        AND osi.dept_code = dm.dept_code
        AND osi.item_code = dm.item_code
        AND osi.sub_code =  dm.sub_code
        AND dm.first_order_date > '20191007'
        and dm.first_order_date = '{0}'
    LEFT JOIN vartefact.service_level_safety2_vinc sl
        on ord.item_code = sl.item_code
        and  ord.sub_code = sl.sub_code
        and  ord.dept_code = sl.dept_code
    JOIN vartefact.forecast_store_item_details id 
        ON ord.item_code = id.item_code
        AND ord.sub_code = id.sub_code
        AND ord.dept_code = id.dept_code
        AND ord.store_code = id.store_code
    LEFT JOIN vartefact.forecast_p4cm_store_item fpsi
        on ord.item_code = fpsi.item_code
        and ord.sub_code = fpsi.sub_code
        and ord.dept_code = fpsi.dept_code 
        and ord.store_code = fpsi.store_code 
        and fpsi.date_key = '{1}'
        """.replace("\n", " ").format(run_date.strftime("%Y%m%d"), stock_date.strftime("%Y%m%d"))

    xdock_df = sqlc.sql(xdock_sql)

    xdock_order = xdock_df.toPandas()

    # +

    xdock_order['order_qty_without_pcb'] = xdock_order['order_without_pcb'].fillna(0)

    xdock_order['dm_order_qty'] = xdock_order['dm_order_qty'].fillna(0)

    xdock_order['dm_order_qty_without_pcb'] = xdock_order['dm_order_qty_without_pcb'].fillna(0)

    xdock_order['four_weeks_after_dm'] = xdock_order['four_weeks_after_dm'].fillna(0)

    xdock_order['service_level'] = xdock_order['service_level'].fillna(1)

    xdock_order['order_qty_with_sl'] = np.round(xdock_order['order_without_pcb'] * (2 - xdock_order['service_level']),
                                                2)

    xdock_order['order_qty'] = np.ceil(xdock_order['order_qty_with_sl'] / xdock_order['qty_per_unit']) * xdock_order[
        'qty_per_unit']

    xdock_order['total_order'] = xdock_order['order_qty'] + xdock_order['dm_order_qty']

    xdock_order['total_order_in_unit'] = xdock_order['total_order'] / xdock_order['qty_per_unit']

    xdock_order['itm_npp'] = xdock_order['itm_npp'].fillna(1)

    xdock_order['order_value'] = xdock_order['itm_npp'] * xdock_order['total_order']

    xdock_order['single_unit_value'] = xdock_order['itm_npp'] * xdock_order['qty_per_unit']
    
    onstock_store.to_csv(record_folder + f'/order_files_debug/onstock_store_{date_str}.csv', header=1)
    
    xdock_order.to_csv(record_folder + f'/order_files_debug/xdock_orders_{date_str}.csv', header=1)

    # +
    wb = Workbook()
    ws = wb.active
    ws.append(['Store_Code', 'Dept_Code', 'Supplier_Code', 'Item_Code',
               'Sub_code', 'Order_Qty', 'Free_Goods_Qty', 'Delv_yyyymmdd',
               'Order_Qty_In_Pieces', 'Order_By', 'Qty_Per_Pack', 'Pack_Per_Box',
               'Regular_Order', 'Regular_Order_Without_PCB', 'DM_Order', 'DM_Order_Without_PCB',
               'PPP', 'NPP', '4_Weeks_After_DM_Order'])

    for index, ord in onstock_store.iterrows():
        ws.append([ord.store_code, ord.dept_code, ord.supplier_code, ord.item_code,
                   ord.sub_code, ord.total_order_in_unit, 0, ord.delivery_day,
                   ord.total_order, ord.order_by, ord.qty_per_pack, ord.pack_per_box,
                   ord.order_qty, ord.order_without_pcb, ord.dm_order_qty, ord.dm_order_qty_without_pcb,
                   ord.ppp, ord.npp, ord.four_weeks_after_dm])

    for index, ord in xdock_order.iterrows():
        ws.append([ord.store_code, ord.dept_code, ord.supplier_code, ord.item_code,
                   ord.sub_code, ord.total_order_in_unit, 0, ord.delivery_day,
                   ord.total_order, ord.order_by, ord.qty_per_pack, ord.pack_per_box,
                   ord.order_qty, ord.order_without_pcb, ord.dm_order_qty, ord.dm_order_qty_without_pcb,
                   ord.ppp, ord.npp, ord.four_weeks_after_dm])

    wb.save(record_folder + '/order_files/' + store_order_filename)

    wb.save(output_path + '/' + store_order_filename)
    
    if len(onstock_store) > 0 or len(xdock_order) > 0 :
    
        store_orders = pd.read_excel(record_folder + '/order_files/' + store_order_filename
                                 , 'Sheet', header=0, dtype=str).fillna('')
    
        sqlc.createDataFrame(store_orders).createOrReplaceTempView("store_orders")
    
        store_orders_sql = """
            INSERT OVERWRITE TABLE vartefact.forecast_store_daily_order_files
            PARTITION (date_key)
            SELECT 
                store_code ,
                dept_code ,
                supplier_code ,
                item_code ,
                sub_code ,
                order_qty ,
                free_goods_qty ,
                delv_yyyymmdd ,
                order_qty_in_pieces ,
                order_by ,
                qty_per_pack ,
                pack_per_box ,
                regular_order ,
                regular_order_without_pcb ,
                dm_order ,
                dm_order_without_pcb ,
                ppp ,
                npp ,
                4_weeks_after_dm_order ,
                {0} as date_key
            FROM store_orders
        """.replace("\n", " ").format(run_date.strftime("%Y%m%d"))

        sqlc.sql(store_orders_sql)

    high_value_onstock_orders = onstock_store[onstock_store['order_value'] >= 2000]

    high_value_onstock_orders = high_value_onstock_orders.sort_values(by='single_unit_value', ascending=True)

    high_value_xdock_orders = xdock_order[xdock_order['order_value'] >= 2000]

    high_value_xdock_orders = high_value_xdock_orders.sort_values(by='single_unit_value', ascending=True)

    wb2 = Workbook()
    ws2 = wb2.active
    ws2.append(['Store_Code', 'Dept_Code', 'Supplier_Code', 'Item_Code',
                'Sub_code', 'Order_Qty_In_Pieces', 'Order_Value', 'Delv_yyyymmdd',
                'Regular_Order', 'Regular_Order_Without_PCB', 'DM_Order', 'DM_Order_Without_PCB',
                '4_Weeks_After_DM_Order', 'Order_Qty_Per_Order_Unit', 'NPP', 'Order_Value_Per_Order_Unit',
                'Item_id', 'Sub_id', 'CN_name'])

    for index, ord in high_value_onstock_orders.iterrows():
        ws2.append([ord.store_code, ord.dept_code, ord.supplier_code, ord.item_code,
                    ord.sub_code, ord.total_order, ord.order_value, ord.delivery_day,
                    ord.order_qty, ord.order_without_pcb, ord.dm_order_qty, ord.dm_order_qty_without_pcb,
                    ord.four_weeks_after_dm, ord.qty_per_unit, ord.itm_npp, ord.single_unit_value,
                    ord.item_id, ord.sub_id, ord.cn_name])

    for index, ord in high_value_xdock_orders.iterrows():
        ws2.append([ord.store_code, ord.dept_code, ord.supplier_code, ord.item_code,
                    ord.sub_code, ord.total_order, ord.order_value, ord.delivery_day,
                    ord.order_qty, ord.order_without_pcb, ord.dm_order_qty, ord.dm_order_qty_without_pcb,
                    ord.four_weeks_after_dm, ord.qty_per_unit, ord.itm_npp, ord.single_unit_value,
                    ord.item_id, ord.sub_id, ord.cn_name])

    wb2.save(record_folder + '/order_checks/' + store_highvalue_order_filename)

    wb2.save(datachecking_output_folder + '/' + store_highvalue_order_filename)

    high_volume_xdock_orders = xdock_order[xdock_order['total_order_in_unit'] > 1]

    high_volume_xdock_orders = high_volume_xdock_orders.sort_values(by='qty_per_unit', ascending=False)

    wb3 = Workbook()
    ws3 = wb3.active
    ws3.append(['Store_Code', 'Dept_Code', 'Supplier_Code', 'Item_Code',
                'Sub_code', 'Order_Qty', 'Order_Qty_In_Pieces', 'Order_Value',
                'Delv_yyyymmdd', 'Regular_Order', 'Regular_Order_With_Service_Level', 'Regular_Order_Without_PCB',
                'Qty_Per_Order_Unit', 'DM_Order', 'DM_Order_Without_PCB', '4_Weeks_After_DM_Order',
                'Service_Level', 'Item_id', 'Sub_id', 'CN_name'])

    for index, ord in high_volume_xdock_orders.iterrows():
        ws3.append([ord.store_code, ord.dept_code, ord.supplier_code, ord.item_code,
                    ord.sub_code, ord.total_order_in_unit, ord.total_order, ord.order_value,
                    ord.delivery_day, ord.order_qty, ord.order_qty_with_sl, ord.order_without_pcb,
                    ord.qty_per_unit, ord.dm_order_qty, ord.dm_order_qty_without_pcb, ord.four_weeks_after_dm,
                    ord.service_level, ord.item_id, ord.sub_id, ord.cn_name])

    wb3.save(record_folder + '/order_checks/' + xdock_high_volume_order_filename)

    wb3.save(datachecking_output_folder + '/' + xdock_high_volume_order_filename)

    sc.stop()


def dc_order_file_process(date_str, record_folder, output_path, dc_order_filename):
    warehouse_location = abspath('spark-warehouse')
    os.environ["PYSPARK_SUBMIT_ARGS"] = '--jars /data/jupyter/kudu-spark2_2.11-1.8.0.jar pyspark-shell'

    spark = SparkSession.builder \
        .appName("Generate DC order file") \
        .config("spark.sql.warehouse.dir", warehouse_location) \
        .config("spark.driver.memory", '8g') \
        .config("spark.executor.memory", '8g') \
        .config("spark.num.executors", '8') \
        .config("hive.exec.compress.output", 'false') \
        .enableHiveSupport() \
        .getOrCreate()

    sc = spark.sparkContext

    sqlc = SQLContext(sc)

    run_date = datetime.datetime.strptime(date_str, '%Y%m%d').date()

    # -
    dc_sql = """
    SELECT ord.dept_code,
        ord.supplier_code,
        ord.item_code,
        ord.sub_code,
        ord.order_qty,
        ord.delivery_day,
        dm.order_qty as dm_order_qty,
        dm.npp,
        cast(sl.service_level AS DOUBLE) service_level,
        cast(dc.qty_per_unit AS float) AS qty_per_unit,
        dc.qty_per_box,
        dc.item_name_english,
        dc.item_name_local,
        dc.current_warehouse,
        dc.primary_barcode
    FROM vartefact.forecast_dc_orders ord
    LEFT JOIN vartefact.forecast_dm_dc_orders dm ON ord.item_id = dm.item_id
        AND ord.sub_id = dm.sub_id
        AND dm.first_order_date > '20191007'
        and dm.first_order_date = '{0}'
    LEFT JOIN vartefact.service_level_safety2_vinc sl ON ord.item_code = sl.item_code
        AND ord.sub_code = sl.sub_code
        AND ord.dept_code = sl.dept_code
    JOIN vartefact.forecast_dc_item_details dc ON ord.item_code = dc.item_code
        AND ord.sub_code = dc.sub_code
        AND ord.dept_code = dc.dept_code
    WHERE ord.order_day = '{0}'
        """.replace("\n", " ")

    dc_df = sqlc.sql(dc_sql.format(run_date.strftime("%Y%m%d")))

    dc_orders = dc_df.toPandas()

    dc_orders['service_level'] = dc_orders['service_level'].fillna(1)

    dc_orders['order_qty'] = dc_orders['order_qty'].fillna(0)

    dc_orders['dm_order_qty'] = dc_orders['dm_order_qty'].fillna(0)

    dc_orders['order_qty_with_sl'] = np.round(dc_orders['order_qty'] * (2 - dc_orders['service_level']), 2)

    dc_orders['order_qty_by_unit'] = np.ceil(
        (dc_orders['order_qty_with_sl'] + dc_orders['dm_order_qty']) / dc_orders['qty_per_box'])
    
    dc_orders['order_in_pieces'] = dc_orders['order_qty_by_unit'] * dc_orders['qty_per_box']

    # +
    wb = Workbook()
    ws = wb.active
    ws.append(
        ['Supplier Code', 'Warehouse', 'Delivery Date', 'Item Code',
         'Item Name', 'ITEM SUBCODE NAME LOCAL', 'POQ quantity', 'Purchase Quantity',
         'Unit', 'Purchase Price', 'Purchase Amount', 'Unit DC Discount',
         'Unit % discount', 'Additional free goods', 'NPP', 'Main barcode',
         'Total order (in Pieces)', 'Service Level', 'Regular Order (in Pieces)', 'DM Order (In Pieces)'])

    for index, ord in dc_orders.iterrows():
        ws.append([ord.dept_code + ord.supplier_code, ord.current_warehouse,
                   ord.delivery_day, ord.dept_code + ord.item_code + ord.sub_code,
                   ord.item_name_english, ord.item_name_local, '', ord.order_qty_by_unit,
                   'B', '', '', '',
                   '', '', ord.npp, ord.primary_barcode,
                   ord.order_in_pieces, ord.service_level, ord.order_qty, ord.dm_order_qty, ])

    wb.save(record_folder + '/order_files/' + dc_order_filename)

    wb.save(output_path + '/' + dc_order_filename)
    
    if len(dc_orders) > 0 :
        dc_orders = pd.read_excel(record_folder + '/order_files/' + dc_order_filename
                                     , 'Sheet', header=0, dtype=str).fillna('')

        dc_orders.columns = ["supplier_code", "warehouse", "delivery_date", "item_code",
                            "item_name", "item_subcode_name_local", "poq_quantity", "purchase_quantity",
                            "unit", "purchase_price", "purchase_amount", "unit_dc_discount",
                            "unit_percent_discount", "additional_free_goods", "npp", "main_barcode",
                            "order_in_pieces", "service_level", "regualr_order_in_pieces", "dm_order_in_pieces"]

        sqlc.createDataFrame(dc_orders).createOrReplaceTempView("dc_orders")

        dc_orders_sql = """
            INSERT OVERWRITE TABLE vartefact.forecast_dc_daily_order_files
            PARTITION (date_key)
            SELECT 
                supplier_code ,
                warehouse ,
                delivery_date ,
                item_code ,

                item_name ,
                item_subcode_name_local ,
                poq_quantity ,
                purchase_quantity ,

                unit ,
                purchase_price ,
                purchase_amount ,
                unit_dc_discount ,

                unit_percent_discount ,
                additional_free_goods ,
                npp ,
                main_barcode ,

                order_in_pieces,
                service_level ,
                regualr_order_in_pieces ,
                dm_order_in_pieces ,
                {0} as date_key
            FROM dc_orders
        """.replace("\n", " ").format(run_date.strftime("%Y%m%d"))

        sqlc.sql(dc_orders_sql)
    
    # -

    sc.stop()
